from fastapi import APIRouter, Depends, HTTPException, status, Query, Path
from sqlalchemy.ext.asyncio import AsyncSession
from uuid import UUID
from typing import Optional, List, Dict, Any
from datetime import datetime

from app.schemas.opportunity import (
    OpportunityCreate,
    OpportunityUpdate,
    OpportunityResponse,
    OpportunityListResponse,
    OpportunityStageUpdate,
    OpportunitySearchRequest,
    OpportunitySearchResult,
    OpportunitySearchResponse,
    OpportunityAnalytics,
    OpportunityInsightsResponse,
    OpportunityPipelineResponse,
    OpportunityForecastResponse
)
from app.services.opportunity import OpportunityService
from app.dependencies.user_auth import get_current_user
from app.dependencies.permissions import get_user_permission
from app.models.user import User
from app.models.opportunity import Opportunity
from app.schemas.user_permission import UserPermissionResponse
from app.db.session import get_request_transaction
from app.utils.logger import get_logger

logger = get_logger("opportunity_routes")

router = APIRouter(prefix="/opportunities", tags=["Opportunities"])

@router.post("/", response_model=OpportunityResponse, status_code=status.HTTP_201_CREATED)
async def create_opportunity(
    opportunity_data: OpportunityCreate,
    db: AsyncSession = Depends(get_request_transaction),
    current_user: User = Depends(get_current_user),
    user_permission: UserPermissionResponse = Depends(get_user_permission({"opportunities": ["create"]}))
) -> OpportunityResponse:
    
    service = OpportunityService(db)
    return await service.create_opportunity(opportunity_data, current_user)

@router.get("/", response_model=OpportunityListResponse)
async def list_opportunities(
    page: int = Query(1, ge=1, description="Page number"),
    size: int = Query(10, ge=1, le=100, description="Page size"),
    stage: Optional[str] = Query(None, description="Filter by stage"),
    search: Optional[str] = Query(None, description="Search query"),
    sort_by: str = Query("created_at", description="Sort field"),
    sort_order: str = Query("desc", pattern="^(asc|desc)$", description="Sort order"),
    market_sector: Optional[str] = Query(None, description="Filter by market sector"),
    state: Optional[str] = Query(None, description="Filter by state"),
    min_value: Optional[float] = Query(None, description="Minimum project value"),
    max_value: Optional[float] = Query(None, description="Maximum project value"),
    risk_level: Optional[str] = Query(None, description="Filter by risk level"),
    min_match_score: Optional[int] = Query(None, description="Minimum match score"),
    account_id: Optional[UUID] = Query(None, description="Filter by account ID"),
    db: AsyncSession = Depends(get_request_transaction),
    current_user: User = Depends(get_current_user),
    user_permission: UserPermissionResponse = Depends(get_user_permission({"opportunities": ["view"]}))
) -> OpportunityListResponse:
    
    stage_enum = None
    if stage:
        try:
            from app.models.opportunity import OpportunityStage
            stage_enum = OpportunityStage(stage)
        except ValueError:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail=f"Invalid stage: {stage}"
            )
    
    service = OpportunityService(db)
    return await service.list_opportunities(
        user=current_user,
        page=page,
        size=size,
        stage=stage_enum,
        search=search,
        sort_by=sort_by,
        sort_order=sort_order,
        market_sector=market_sector,
        state=state,
        min_value=min_value,
        max_value=max_value,
        risk_level=risk_level,
        min_match_score=min_match_score,
        account_id=account_id,
    )

@router.get("/{opportunity_id}", response_model=OpportunityResponse)
async def get_opportunity(
    opportunity_id: str = Path(..., description="Opportunity ID (UUID or custom ID like OPP-NY0001)"),
    db: AsyncSession = Depends(get_request_transaction),
    current_user: User = Depends(get_current_user),
    user_permission: UserPermissionResponse = Depends(get_user_permission({"opportunities": ["view"]}))
) -> OpportunityResponse:
    
    service = OpportunityService(db)
    
    if opportunity_id.startswith('OPP-NY'):
        opportunity = await service.get_opportunity_by_custom_id(opportunity_id, current_user)
    else:
        try:
            uuid_id = UUID(opportunity_id)
            opportunity = await service.get_opportunity_by_id(uuid_id, current_user)
        except ValueError:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Invalid opportunity ID format"
            )
    
    if not opportunity:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Opportunity not found"
        )
    
    return opportunity

@router.put("/{opportunity_id}", response_model=OpportunityResponse)
async def update_opportunity(
    opportunity_id: UUID = Path(..., description="Opportunity ID"),
    opportunity_data: OpportunityUpdate = None,
    db: AsyncSession = Depends(get_request_transaction),
    current_user: User = Depends(get_current_user),
    user_permission: UserPermissionResponse = Depends(get_user_permission({"opportunities": ["edit"]}))
) -> OpportunityResponse:
    
    service = OpportunityService(db)
    opportunity = await service.update_opportunity(opportunity_id, opportunity_data, current_user)
    
    if not opportunity:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Opportunity not found"
        )
    
    return opportunity

@router.delete("/{opportunity_id}", status_code=status.HTTP_204_NO_CONTENT)
async def delete_opportunity(
    opportunity_id: UUID = Path(..., description="Opportunity ID"),
    db: AsyncSession = Depends(get_request_transaction),
    current_user: User = Depends(get_current_user),
    user_permission: UserPermissionResponse = Depends(get_user_permission({"opportunities": ["delete"]}))
) -> None:
    
    service = OpportunityService(db)
    success = await service.delete_opportunity(opportunity_id, current_user)
    
    if not success:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Opportunity not found"
        )

@router.put("/{opportunity_id}/stage", response_model=OpportunityResponse)
async def update_opportunity_stage(
    opportunity_id: UUID = Path(..., description="Opportunity ID"),
    stage_data: OpportunityStageUpdate = None,
    db: AsyncSession = Depends(get_request_transaction),
    current_user: User = Depends(get_current_user),
    user_permission: UserPermissionResponse = Depends(get_user_permission({"opportunities": ["edit"]}))
) -> OpportunityResponse:
    
    service = OpportunityService(db)
    opportunity = await service.update_opportunity_stage(opportunity_id, stage_data, current_user)
    
    if not opportunity:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Opportunity not found"
        )
    
    return opportunity

@router.get("/analytics/dashboard", response_model=OpportunityAnalytics)
async def get_opportunity_analytics(
    days: int = Query(30, ge=1, le=365, description="Number of days to analyze"),
    db: AsyncSession = Depends(get_request_transaction),
    current_user: User = Depends(get_current_user),
    user_permission: UserPermissionResponse = Depends(get_user_permission({"opportunities": ["view"]}))
) -> OpportunityAnalytics:
    
    service = OpportunityService(db)
    return await service.get_opportunity_analytics(current_user, days)

@router.get("/pipeline/view", response_model=OpportunityPipelineResponse)
async def get_opportunity_pipeline(
    db: AsyncSession = Depends(get_request_transaction),
    current_user: User = Depends(get_current_user),
    user_permission: UserPermissionResponse = Depends(get_user_permission({"opportunities": ["view"]}))
) -> OpportunityPipelineResponse:
    
    service = OpportunityService(db)
    return await service.get_opportunity_pipeline(current_user)

@router.post("/search/ai", response_model=List[OpportunitySearchResult])
async def search_opportunities_ai(
    search_request: OpportunitySearchRequest,
    db: AsyncSession = Depends(get_request_transaction),
    current_user: User = Depends(get_current_user),
    user_permission: UserPermissionResponse = Depends(get_user_permission({"opportunities": ["view"]}))
) -> List[OpportunitySearchResult]:
    
    service = OpportunityService(db)
    results = await service.search_opportunities_ai(search_request, current_user)
    
    return results

@router.get("/{opportunity_id}/insights", response_model=OpportunityInsightsResponse)
async def get_opportunity_insights(
    opportunity_id: UUID = Path(..., description="Opportunity ID"),
    db: AsyncSession = Depends(get_request_transaction),
    current_user: User = Depends(get_current_user),
    user_permission: UserPermissionResponse = Depends(get_user_permission({"opportunities": ["view"]}))
) -> OpportunityInsightsResponse:
    
    service = OpportunityService(db)
    return await service.generate_opportunity_insights(opportunity_id, current_user)

@router.get("/{opportunity_id}/forecast", response_model=OpportunityForecastResponse)
async def get_opportunity_forecast(
    opportunity_id: UUID = Path(..., description="Opportunity ID"),
    period: str = Query("quarterly", pattern="^(monthly|quarterly|yearly)$", description="Forecast period"),
    db: AsyncSession = Depends(get_request_transaction),
    current_user: User = Depends(get_current_user),
    user_permission: UserPermissionResponse = Depends(get_user_permission({"opportunities": ["view"]}))
) -> OpportunityForecastResponse:
    
    from app.schemas.opportunity import OpportunityForecast
    from datetime import datetime, timedelta
    
    forecast = OpportunityForecast(
        period=period,
        forecasted_revenue=100000.0,  # Placeholder
        confidence_level=75.0,  # Placeholder
        scenarios={
            "best_case": 150000.0,
            "worst_case": 75000.0,
            "most_likely": 100000.0
        },
        factors=[
            "Market conditions",
            "Competition level",
            "Client budget",
            "Timeline constraints"
        ]
    )
    
    return OpportunityForecastResponse(
        opportunities=[opportunity_id],
        forecast=forecast,
        generated_at=datetime.utcnow(),
        next_review_date=datetime.utcnow() + timedelta(days=30)
    )

@router.get("/export/csv")
async def export_opportunities_csv(
    stage: Optional[str] = Query(None, description="Filter by stage"),
    market_sector: Optional[str] = Query(None, description="Filter by market sector"),
    state: Optional[str] = Query(None, description="Filter by state"),
    min_value: Optional[float] = Query(None, description="Minimum project value"),
    max_value: Optional[float] = Query(None, description="Maximum project value"),
    risk_level: Optional[str] = Query(None, description="Filter by risk level"),
    db: AsyncSession = Depends(get_request_transaction),
    current_user: User = Depends(get_current_user),
    user_permission: UserPermissionResponse = Depends(get_user_permission({"opportunities": ["view"]}))
):
    """Export opportunities to CSV format."""
    from fastapi.responses import Response
    import csv
    import io
    from datetime import datetime
    
    service = OpportunityService(db)
    
    stage_enum = None
    if stage:
        try:
            from app.models.opportunity import OpportunityStage
            stage_enum = OpportunityStage(stage)
        except ValueError:
            pass
    
    opportunities_list = await service.list_opportunities(
        user=current_user,
        page=1,
        size=10000,  # Large number to get all
        stage=stage_enum,
        market_sector=market_sector,
        state=state,
        min_value=min_value,
        max_value=max_value,
        risk_level=risk_level,
    )
    
    # Generate CSV
    output = io.StringIO()
    writer = csv.writer(output)
    
    # Write headers
    headers = [
        'ID', 'Project Name', 'Client Name', 'Project Value', 'Stage', 'Market Sector',
        'State', 'Location', 'Deadline', 'Match Score', 'Risk Level', 'Status',
        'Created At', 'Updated At'
    ]
    writer.writerow(headers)
    
    # Write data
    for opp in opportunities_list.opportunities:
        writer.writerow([
            str(opp.id),
            opp.project_name or '',
            opp.client_name or '',
            opp.project_value or 0,
            opp.stage.value if opp.stage else '',
            opp.market_sector or '',
            opp.state or '',
            opp.location or '',
            opp.deadline.isoformat() if opp.deadline else '',
            opp.match_score or 0,
            opp.risk_level.value if opp.risk_level else '',
            opp.status.value if opp.status else '',
            opp.created_at.isoformat() if opp.created_at else '',
            opp.updated_at.isoformat() if opp.updated_at else '',
        ])
    
    csv_content = output.getvalue()
    output.close()
    
    filename = f"opportunities_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
    
    return Response(
        content=csv_content,
        media_type="text/csv",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )

@router.get("/export/excel")
async def export_opportunities_excel(
    stage: Optional[str] = Query(None, description="Filter by stage"),
    market_sector: Optional[str] = Query(None, description="Filter by market sector"),
    state: Optional[str] = Query(None, description="Filter by state"),
    min_value: Optional[float] = Query(None, description="Minimum project value"),
    max_value: Optional[float] = Query(None, description="Maximum project value"),
    risk_level: Optional[str] = Query(None, description="Filter by risk level"),
    db: AsyncSession = Depends(get_request_transaction),
    current_user: User = Depends(get_current_user),
    user_permission: UserPermissionResponse = Depends(get_user_permission({"opportunities": ["view"]}))
):
    """Export opportunities to Excel format."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from fastapi.responses import FileResponse
        import tempfile
        from datetime import datetime
        
        service = OpportunityService(db)
        
        stage_enum = None
        if stage:
            try:
                from app.models.opportunity import OpportunityStage
                stage_enum = OpportunityStage(stage)
            except ValueError:
                pass
        
        opportunities_list = await service.list_opportunities(
            user=current_user,
            page=1,
            size=10000,
            stage=stage_enum,
            market_sector=market_sector,
            state=state,
            min_value=min_value,
            max_value=max_value,
            risk_level=risk_level,
        )
        
        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Opportunities"
        
        # Headers
        headers = [
            'ID', 'Project Name', 'Client Name', 'Project Value', 'Stage', 'Market Sector',
            'State', 'Location', 'Deadline', 'Match Score', 'Risk Level', 'Status',
            'Created At', 'Updated At'
        ]
        
        # Style header row
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Write data
        for row_idx, opp in enumerate(opportunities_list.opportunities, 2):
            ws.cell(row=row_idx, column=1, value=str(opp.id))
            ws.cell(row=row_idx, column=2, value=opp.project_name or '')
            ws.cell(row=row_idx, column=3, value=opp.client_name or '')
            ws.cell(row=row_idx, column=4, value=opp.project_value or 0)
            ws.cell(row=row_idx, column=5, value=opp.stage.value if opp.stage else '')
            ws.cell(row=row_idx, column=6, value=opp.market_sector or '')
            ws.cell(row=row_idx, column=7, value=opp.state or '')
            ws.cell(row=row_idx, column=8, value=opp.location or '')
            ws.cell(row=row_idx, column=9, value=opp.deadline.isoformat() if opp.deadline else '')
            ws.cell(row=row_idx, column=10, value=opp.match_score or 0)
            ws.cell(row=row_idx, column=11, value=opp.risk_level.value if opp.risk_level else '')
            ws.cell(row=row_idx, column=12, value=opp.status.value if opp.status else '')
            ws.cell(row=row_idx, column=13, value=opp.created_at.isoformat() if opp.created_at else '')
            ws.cell(row=row_idx, column=14, value=opp.updated_at.isoformat() if opp.updated_at else '')
        
        # Auto-adjust column widths
        for col in ws.columns:
            max_length = 0
            col_letter = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[col_letter].width = adjusted_width
        
        filename = f"opportunities_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
        temp_file.close()
        wb.save(temp_file.name)
        
        return FileResponse(
            path=temp_file.name,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            filename=filename,
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
    except ImportError:
        raise HTTPException(
            status_code=500,
            detail="Excel export requires openpyxl. Install with: pip install openpyxl"
        )

@router.get("/export/pdf")
async def export_opportunities_pdf(
    stage: Optional[str] = Query(None, description="Filter by stage"),
    market_sector: Optional[str] = Query(None, description="Filter by market sector"),
    state: Optional[str] = Query(None, description="Filter by state"),
    min_value: Optional[float] = Query(None, description="Minimum project value"),
    max_value: Optional[float] = Query(None, description="Maximum project value"),
    risk_level: Optional[str] = Query(None, description="Filter by risk level"),
    db: AsyncSession = Depends(get_request_transaction),
    current_user: User = Depends(get_current_user),
    user_permission: UserPermissionResponse = Depends(get_user_permission({"opportunities": ["view"]}))
):
    """Export opportunities to PDF format."""
    try:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import letter, A4
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import inch
        from fastapi.responses import Response
        import io
        from datetime import datetime
        
        service = OpportunityService(db)
        
        stage_enum = None
        if stage:
            try:
                from app.models.opportunity import OpportunityStage
                stage_enum = OpportunityStage(stage)
            except ValueError:
                pass
        
        opportunities_list = await service.list_opportunities(
            user=current_user,
            page=1,
            size=1000,  # Limit for PDF
            stage=stage_enum,
            market_sector=market_sector,
            state=state,
            min_value=min_value,
            max_value=max_value,
            risk_level=risk_level,
        )
        
        # Create PDF in memory
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4, topMargin=0.5*inch, bottomMargin=0.5*inch)
        
        # Container for PDF elements
        elements = []
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=18,
            textColor=colors.HexColor('#161950'),
            spaceAfter=30,
        )
        
        # Title
        elements.append(Paragraph("Opportunities Export Report", title_style))
        elements.append(Spacer(1, 0.2*inch))
        
        # Summary
        summary_text = f"Total Opportunities: {opportunities_list.total} | Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        elements.append(Paragraph(summary_text, styles['Normal']))
        elements.append(Spacer(1, 0.3*inch))
        
        # Prepare table data
        data = [['Project Name', 'Client', 'Value', 'Stage', 'Sector', 'State']]
        
        for opp in opportunities_list.opportunities[:100]:  # Limit to 100 for PDF
            data.append([
                opp.project_name or 'N/A',
                opp.client_name or 'N/A',
                f"${opp.project_value:,.0f}" if opp.project_value else 'N/A',
                opp.stage.value if opp.stage else 'N/A',
                opp.market_sector or 'N/A',
                opp.state or 'N/A',
            ])
        
        # Create table
        table = Table(data, colWidths=[1.5*inch, 1.2*inch, 1*inch, 0.8*inch, 1*inch, 0.8*inch])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#366092')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.grey),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey]),
        ]))
        
        elements.append(table)
        
        # Build PDF
        doc.build(elements)
        buffer.seek(0)
        
        filename = f"opportunities_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
        
        return Response(
            content=buffer.getvalue(),
            media_type="application/pdf",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
    except ImportError:
        raise HTTPException(
            status_code=500,
            detail="PDF export requires reportlab. Install with: pip install reportlab"
        )

@router.get("/by-account/{account_id}", response_model=OpportunityListResponse)
async def list_opportunities_by_account(
    account_id: UUID = Path(..., description="Account ID"),
    page: int = Query(1, ge=1, description="Page number"),
    size: int = Query(10, ge=1, le=100, description="Page size"),
    stage: Optional[str] = Query(None, description="Filter by stage"),
    db: AsyncSession = Depends(get_request_transaction),
    current_user: User = Depends(get_current_user),
    user_permission: UserPermissionResponse = Depends(get_user_permission({"opportunities": ["view"]}))
) -> OpportunityListResponse:
    
    stage_enum = None
    if stage:
        try:
            from app.models.opportunity import OpportunityStage
            stage_enum = OpportunityStage(stage)
        except ValueError:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail=f"Invalid stage: {stage}"
            )
    
    offset = (page - 1) * size
    
    from sqlalchemy import select, func
    query = select(Opportunity).where(
        Opportunity.org_id == current_user.org_id,
        Opportunity.account_id == account_id
    )
    
    if stage_enum:
        query = query.where(Opportunity.stage == stage_enum)
    
    query = query.order_by(Opportunity.created_at.desc()).limit(size).offset(offset)
    result = await db.execute(query)
    opportunities = result.scalars().all()
    
    total_stmt = select(func.count(Opportunity.id)).where(
        Opportunity.org_id == current_user.org_id,
        Opportunity.account_id == account_id
    )
    
    if stage_enum:
        total_stmt = total_stmt.where(Opportunity.stage == stage_enum)
    
    total_result = await db.execute(total_stmt)
    total = total_result.scalar() or 0
    
    opportunity_responses = [OpportunityResponse.model_validate(opp) for opp in opportunities]
    
    return OpportunityListResponse(
        opportunities=opportunity_responses,
        total=total,
        page=page,
        size=size,
        total_pages=(total + size - 1) // size if total > 0 else 0
    )

@router.get("/health/check")
async def health_check():
    
    return {
        "status": "healthy",
        "module": "opportunities",
        "timestamp": "2024-01-01T00:00:00Z"
    }

@router.post("/{opportunity_id}/ai-analysis/comprehensive")
async def perform_comprehensive_ai_analysis(
    opportunity_id: UUID = Path(..., description="Opportunity ID"),
    db: AsyncSession = Depends(get_request_transaction),
    current_user: User = Depends(get_current_user),
    user_permission: UserPermissionResponse = Depends(get_user_permission({"opportunities": ["view"]}))
) -> Dict[str, Any]:
    """Perform comprehensive AI analysis for an opportunity."""
    from app.services.opportunity_ai_analysis import OpportunityAIAnalysisService
    
    service = OpportunityAIAnalysisService(db)
    analysis = await service.perform_comprehensive_analysis(opportunity_id)
    
    if "error" in analysis:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND if "not found" in analysis["error"].lower() else status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=analysis["error"]
        )
    
    return analysis

@router.post("/{opportunity_id}/ai-analysis/competition")
async def analyze_competition(
    opportunity_id: UUID = Path(..., description="Opportunity ID"),
    db: AsyncSession = Depends(get_request_transaction),
    current_user: User = Depends(get_current_user),
    user_permission: UserPermissionResponse = Depends(get_user_permission({"opportunities": ["view"]}))
) -> Dict[str, Any]:
    """Analyze competition for an opportunity."""
    from app.services.opportunity_ai_analysis import OpportunityAIAnalysisService
    from app.models.opportunity import Opportunity
    from app.models.opportunity_tabs import OpportunityOverview
    from sqlalchemy import select
    
    opportunity = await Opportunity.get_by_id(opportunity_id, current_user.org_id)
    if not opportunity:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Opportunity not found")
    
    stmt = select(OpportunityOverview).where(OpportunityOverview.opportunity_id == opportunity_id)
    result = await db.execute(stmt)
    overview = result.scalar_one_or_none()
    
    service = OpportunityAIAnalysisService(db)
    analysis = await service.analyze_competition(opportunity, overview)
    return analysis

@router.post("/{opportunity_id}/ai-analysis/technical")
async def analyze_technical_fit(
    opportunity_id: UUID = Path(..., description="Opportunity ID"),
    db: AsyncSession = Depends(get_request_transaction),
    current_user: User = Depends(get_current_user),
    user_permission: UserPermissionResponse = Depends(get_user_permission({"opportunities": ["view"]}))
) -> Dict[str, Any]:
    """Analyze technical fit for an opportunity."""
    from app.services.opportunity_ai_analysis import OpportunityAIAnalysisService
    from app.models.opportunity import Opportunity
    from app.models.opportunity_tabs import OpportunityOverview
    from sqlalchemy import select
    
    opportunity = await Opportunity.get_by_id(opportunity_id, current_user.org_id)
    if not opportunity:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Opportunity not found")
    
    stmt = select(OpportunityOverview).where(OpportunityOverview.opportunity_id == opportunity_id)
    result = await db.execute(stmt)
    overview = result.scalar_one_or_none()
    
    service = OpportunityAIAnalysisService(db)
    analysis = await service.analyze_technical_fit(opportunity, overview)
    return analysis

@router.post("/{opportunity_id}/ai-analysis/financial")
async def analyze_financial_viability(
    opportunity_id: UUID = Path(..., description="Opportunity ID"),
    db: AsyncSession = Depends(get_request_transaction),
    current_user: User = Depends(get_current_user),
    user_permission: UserPermissionResponse = Depends(get_user_permission({"opportunities": ["view"]}))
) -> Dict[str, Any]:
    """Analyze financial viability for an opportunity."""
    from app.services.opportunity_ai_analysis import OpportunityAIAnalysisService
    from app.models.opportunity import Opportunity
    from app.models.opportunity_tabs import OpportunityOverview
    from sqlalchemy import select
    
    opportunity = await Opportunity.get_by_id(opportunity_id, current_user.org_id)
    if not opportunity:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Opportunity not found")
    
    stmt = select(OpportunityOverview).where(OpportunityOverview.opportunity_id == opportunity_id)
    result = await db.execute(stmt)
    overview = result.scalar_one_or_none()
    
    service = OpportunityAIAnalysisService(db)
    analysis = await service.analyze_financial_viability(opportunity, overview)
    return analysis

@router.post("/{opportunity_id}/ai-analysis/recommendations")
async def get_strategic_recommendations(
    opportunity_id: UUID = Path(..., description="Opportunity ID"),
    db: AsyncSession = Depends(get_request_transaction),
    current_user: User = Depends(get_current_user),
    user_permission: UserPermissionResponse = Depends(get_user_permission({"opportunities": ["view"]}))
) -> Dict[str, Any]:
    """Get strategic recommendations for an opportunity."""
    from app.services.opportunity_ai_analysis import OpportunityAIAnalysisService
    from app.models.opportunity import Opportunity
    
    opportunity = await Opportunity.get_by_id(opportunity_id, current_user.org_id)
    if not opportunity:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Opportunity not found")
    
    service = OpportunityAIAnalysisService(db)
    recommendations = await service.generate_strategic_recommendations(opportunity)
    return recommendations

@router.post("/{opportunity_id}/create-proposal")
async def create_proposal_from_opportunity(
    opportunity_id: UUID = Path(..., description="Opportunity ID"),
    db: AsyncSession = Depends(get_request_transaction),
    current_user: User = Depends(get_current_user),
    user_permission: UserPermissionResponse = Depends(get_user_permission({"proposals": ["create"]}))
) -> Dict[str, Any]:
    """Create a proposal from an opportunity."""
    from app.services.proposal import ProposalService
    from app.schemas.proposal import ProposalCreate
    from app.models.opportunity import Opportunity
    from sqlalchemy import select
    
    # Get opportunity
    stmt = select(Opportunity).where(
        Opportunity.id == opportunity_id,
        Opportunity.org_id == current_user.org_id
    )
    result = await db.execute(stmt)
    opportunity = result.scalar_one_or_none()
    
    if not opportunity:
        raise HTTPException(status_code=404, detail="Opportunity not found")
    
    # Create proposal with opportunity data
    proposal_data = ProposalCreate(
        opportunity_id=opportunity_id,
        account_id=opportunity.account_id,
        title=f"Proposal for {opportunity.project_name}",
        summary=opportunity.description,
        total_value=opportunity.project_value,
        currency=opportunity.currency or "USD",
        due_date=opportunity.deadline,
    )
    
    service = ProposalService(db)
    proposal = await service.create_proposal(proposal_data, current_user)
    
    return {
        "proposal_id": str(proposal.id),
        "proposal_number": proposal.proposal_number,
        "message": "Proposal created successfully from opportunity"
    }

@router.post("/{opportunity_id}/create-budget")
async def create_budget_from_opportunity(
    opportunity_id: UUID = Path(..., description="Opportunity ID"),
    budget_year: Optional[str] = Query(None, description="Budget year (defaults to current year)"),
    db: AsyncSession = Depends(get_request_transaction),
    current_user: User = Depends(get_current_user),
    user_permission: UserPermissionResponse = Depends(get_user_permission({"finance": ["create"]}))
) -> Dict[str, Any]:
    """Create a finance budget from opportunity financial data."""
    from app.models.opportunity import Opportunity
    from app.models.opportunity_tabs import OpportunityFinancial
    from app.services.finance_planning import save_annual_budget
    from app.schemas.finance import FinanceAnnualBudgetCreate
    from sqlalchemy import select
    from datetime import datetime
    
    # Get opportunity
    stmt = select(Opportunity).where(
        Opportunity.id == opportunity_id,
        Opportunity.org_id == current_user.org_id
    )
    result = await db.execute(stmt)
    opportunity = result.scalar_one_or_none()
    
    if not opportunity:
        raise HTTPException(status_code=404, detail="Opportunity not found")
    
    # Get financial summary if available
    financial_stmt = select(OpportunityFinancial).where(
        OpportunityFinancial.opportunity_id == opportunity_id
    )
    financial_result = await db.execute(financial_stmt)
    financial = financial_result.scalar_one_or_none()
    
    # Determine budget year
    if not budget_year:
        budget_year = str(datetime.now().year)
    
    # Build expense lines from opportunity financial data
    expense_lines = []
    if financial and financial.budget_categories:
        for category in financial.budget_categories:
            if isinstance(category, dict) and category.get("amount"):
                expense_lines.append({
                    "label": category.get("name", "Category"),
                    "target": float(category.get("amount", 0)),
                    "variance": 0.0
                })
    
    # If no categories, use project value as single expense line
    if not expense_lines and opportunity.project_value:
        expense_lines.append({
            "label": f"Project: {opportunity.project_name}",
            "target": float(opportunity.project_value),
            "variance": 0.0
        })
    
    # Create budget
    budget_data = FinanceAnnualBudgetCreate(
        budget_year=budget_year,
        target_growth_rate=15.0,
        total_revenue_target=0.0,
        total_expense_budget=sum(line["target"] for line in expense_lines),
        revenue_lines=[],
        expense_lines=expense_lines,
    )
    
    budget = await save_annual_budget(
        db=db,
        budget_data=budget_data,
        user_id=current_user.id,
        org_id=current_user.org_id
    )
    
    return {
        "budget_id": str(budget.id),
        "budget_year": budget.budget_year,
        "total_expense_budget": float(budget.total_expense_budget),
        "message": "Budget created successfully from opportunity"
    }