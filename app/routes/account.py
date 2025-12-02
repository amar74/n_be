from fastapi import APIRouter, Depends, Query, Path, HTTPException, Response, Request
from starlette.requests import Request
from fastapi.responses import StreamingResponse, FileResponse
from typing import Optional, List
from uuid import UUID
from datetime import datetime
from sqlalchemy import select, and_, func
from sqlalchemy.ext.asyncio import AsyncSession
import csv
import io
import tempfile
import os

from app.schemas.account import (
    AccountCreate, AccountListResponse, AccountListItem, AccountDetailResponse, AccountUpdate, 
    ContactCreate, ContactResponse, ContactListResponse, AddressResponse,
    ContactAddRequest, ContactUpdateRequest,
    AccountCreateResponse, AccountUpdateResponse, AccountDeleteResponse,
    ContactCreateResponse, ContactUpdateResponse, ContactDeleteResponse,
    AccountApprovalRequest
)
from app.services.account import account_service
from app.dependencies.user_auth import get_current_user
from app.dependencies.permissions import get_user_permission
from app.models.user import User
from app.models.account import Account
from app.schemas.user_permission import UserPermissionResponse
from app.schemas.auth import AuthUserResponse
from app.utils.logger import logger
from app.db.session import get_request_transaction

router = APIRouter(prefix="/accounts", tags=["accounts"])


# ==================== ACCOUNT ACTIVITIES ====================

@router.get("/{account_id}/activities")
async def get_account_activities(
    account_id: str,
    limit: int = Query(10, ge=1, le=50),
    current_user: AuthUserResponse = Depends(get_current_user)
):
    """
    Get recent activities for a specific account
    Returns: notes, documents, opportunities, team changes, contact changes
    """
    try:
        from app.models.account_note import AccountNote
        from app.models.account_document import AccountDocument
        from app.models.opportunity import Opportunity
        from app.models.account_team import AccountTeam
        from app.models.contact import Contact
        
        db = get_request_transaction()
        activities = []
        
        # Fetch notes
        notes_result = await db.execute(
            select(AccountNote)
            .where(AccountNote.account_id == UUID(account_id))
            .order_by(AccountNote.created_at.desc())
            .limit(limit)
        )
        notes = notes_result.scalars().all()
        for note in notes:
            activities.append({
                'id': str(note.id),
                'type': 'note',
                'title': f"Note added: {note.title}",
                'timestamp': note.created_at.isoformat(),
                'color': '#2563EB'  # Blue
            })
        
        # Fetch documents
        docs_result = await db.execute(
            select(AccountDocument)
            .where(AccountDocument.account_id == UUID(account_id))
            .order_by(AccountDocument.created_at.desc())
            .limit(limit)
        )
        docs = docs_result.scalars().all()
        for doc in docs:
            activities.append({
                'id': str(doc.id),
                'type': 'document',
                'title': f"Document uploaded: {doc.name}",
                'timestamp': doc.created_at.isoformat(),
                'color': '#9333EA'  # Purple
            })
        
        # Fetch opportunities
        opps_result = await db.execute(
            select(Opportunity)
            .where(Opportunity.account_id == UUID(account_id))
            .order_by(Opportunity.created_at.desc())
            .limit(limit)
        )
        opps = opps_result.scalars().all()
        for opp in opps:
            activities.append({
                'id': str(opp.id),
                'type': 'opportunity',
                'title': f"Opportunity created: {opp.project_name}",
                'timestamp': opp.created_at.isoformat(),
                'color': '#16A34A'  # Green
            })
        
        # Fetch team assignments
        team_result = await db.execute(
            select(AccountTeam)
            .where(AccountTeam.account_id == UUID(account_id))
            .order_by(AccountTeam.assigned_at.desc())
            .limit(limit)
        )
        team = team_result.scalars().all()
        for member in team:
            activities.append({
                'id': str(member.id),
                'type': 'team',
                'title': f"Team member assigned" + (f": {member.role_in_account}" if member.role_in_account else ""),
                'timestamp': member.assigned_at.isoformat(),
                'color': '#EAB308'  # Yellow
            })
        
        # Fetch contacts
        contacts_result = await db.execute(
            select(Contact)
            .where(Contact.account_id == UUID(account_id))
            .order_by(Contact.created_at.desc())
            .limit(limit)
        )
        contacts = contacts_result.scalars().all()
        for contact in contacts:
            activities.append({
                'id': str(contact.id),  # Fixed: Contact.id not contact_id
                'type': 'contact',
                'title': f"Contact added: {contact.name}",
                'timestamp': contact.created_at.isoformat(),
                'color': '#06B6D4'  # Cyan
            })
        
        # Sort all activities by timestamp (most recent first)
        activities.sort(key=lambda x: x['timestamp'], reverse=True)
        
        # Return only the most recent N activities
        return {'activities': activities[:limit]}
        
    except Exception as e:
        logger.error(f"Error fetching account activities: {e}")
        raise HTTPException(
            status_code=500,
            detail=f"Failed to fetch account activities: {str(e)}"
        )

# Get accounts for the current user's organization
@router.get("/", response_model=AccountListResponse)
async def list_accounts(
    page: int = Query(1, ge=1),
    size: int = Query(10, ge=1, le=100),
    current_user: AuthUserResponse = Depends(get_current_user)
):
    try:
        accounts = await account_service.get_accounts_for_organization(current_user.org_id)
        
        # Simple pagination
        start_idx = (page - 1) * size
        end_idx = start_idx + size
        paginated_accounts = accounts[start_idx:end_idx]
        
        # Convert to response format
        account_items = []
        for account in paginated_accounts:
            # Build primary contact data if available
            primary_contact_name = None
            primary_contact_email = None
            try:
                if account.primary_contact:
                    primary_contact_name = account.primary_contact.name
                    primary_contact_email = account.primary_contact.email
            except Exception as e:
                logger.error(f"Error accessing primary_contact: {str(e)}")
            
            # Build address data if available
            client_address = None
            try:
                if account.client_address:
                    logger.info(f"Account {account.account_id} has address: {account.client_address.city}, {account.client_address.line1}")
                    client_address = AddressResponse(
                        address_id=account.client_address.id,
                        line1=account.client_address.line1,
                        line2=account.client_address.line2,
                        city=account.client_address.city,
                        state=account.client_address.state,
                        pincode=account.client_address.pincode
                    )
                else:
                    logger.info(f"Account {account.account_id} has no address data")
            except Exception as e:
                logger.error(f"Error creating AddressResponse: {str(e)}")
                client_address = None
            
            # Get approval status from database or default to pending
            approval_status = getattr(account, 'approval_status', None) or "pending"
            approval_notes = getattr(account, 'approval_notes', None)
            
            # Get creator name if available
            created_by_name = None
            try:
                if hasattr(account, 'creator') and account.creator:
                    created_by_name = account.creator.name or account.creator.email
            except Exception as e:
                logger.error(f"Error accessing creator: {str(e)}")
            
            account_items.append(AccountListItem(
                account_id=account.account_id,
                custom_id=account.custom_id,
                client_name=account.client_name,
                client_address=client_address,
                primary_contact_name=primary_contact_name,
                primary_contact_email=primary_contact_email,
                client_type=account.client_type,
                market_sector=account.market_sector,
                total_value=float(account.total_value) if account.total_value else None,
                ai_health_score=float(account.ai_health_score) if account.ai_health_score else None,
                health_trend=account.health_trend,
                risk_level=account.risk_level,
                last_contact=account.last_contact,
                approval_status=approval_status,
                account_approver=account.account_approver,
                approval_date=account.approval_date,
                created_by_name=created_by_name,
                created_at=account.created_at
            ))
        
        return AccountListResponse(
            accounts=account_items,
            pagination={
                "page": page,
                "size": size,
                "total": len(accounts),
                "total_pages": (len(accounts) + size - 1) // size
            }
        )
        
    except Exception as e:
        logger.error(f"Error listing accounts: {str(e).replace('{', '{{').replace('}', '}}')}", exc_info=True)
        raise HTTPException(status_code=500, detail="Failed to retrieve accounts")

# Get single account by ID
# Export accounts to CSV, Excel, or PDF
@router.get("/export")
async def export_accounts(
    format: str = Query("csv", regex="^(csv|excel|pdf)$", description="Export format: csv, excel, or pdf"),
    current_user: AuthUserResponse = Depends(get_current_user)
):
    """
    Export all accounts for the current user's organization to CSV, Excel, or PDF format
    """
    try:
        logger.info(f"Export request received from user {current_user.id}")
        
        # Validate org_id
        if not current_user.org_id:
            raise HTTPException(status_code=400, detail="User is not associated with an organization")
        
        # Get all accounts for the organization
        accounts = await account_service.get_accounts_for_organization(current_user.org_id)
        
        if not accounts:
            raise HTTPException(status_code=404, detail="No accounts found to export")
        
        # Create CSV content
        output = io.StringIO()
        writer = csv.writer(output)
        
        # Write headers
        headers = [
            'Account Name',
            'City', 
            'Hosting Area',
            'Type',
            'Contact',
            'Tier Type',
            'Health Score',
            'Risk',
            'Total Value',
            'Status',
            'Created Date',
            'Email',
            'Phone'
        ]
        writer.writerow(headers)
        
        # Write account data
        for account in accounts:
            # Get risk level based on health score
            health_score = account.ai_health_score or 0
            if health_score >= 80:
                risk = 'Low'
            elif health_score >= 50:
                risk = 'Medium'
            else:
                risk = 'High'
            
            # Get primary contact info
            primary_contact_name = None
            primary_contact_email = None
            primary_contact_phone = None
            try:
                if account.primary_contact:
                    primary_contact_name = account.primary_contact.name
                    primary_contact_email = account.primary_contact.email
                    primary_contact_phone = account.primary_contact.phone
            except Exception:
                pass
            
            # Get approval status (derived from account_approver and approval_date)
            approval_status = "pending"
            if account.account_approver and account.approval_date:
                approval_status = "approved"
            
            writer.writerow([
                account.client_name or '',
                account.client_address.city if account.client_address else '',
                account.hosting_area or '',
                account.market_sector or '',
                primary_contact_name or '',
                account.client_type.value if account.client_type else '',
                f"{health_score}%",
                risk,
                f"${(account.total_value or 0):.1f}M",
                approval_status,
                account.created_at.strftime('%Y-%m-%d') if account.created_at else '',
                primary_contact_email or '',
                primary_contact_phone or ''
            ])
        
        # Get CSV content
        output.seek(0)
        csv_content = output.getvalue()
        output.close()
        
        # Create temporary file based on format
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        
        if format == "excel":
            # Excel export (requires openpyxl - install with: pip install openpyxl)
            try:
                from openpyxl import Workbook
                from openpyxl.styles import Font, Alignment, PatternFill
                
                wb = Workbook()
                ws = wb.active
                ws.title = "Accounts"
                
                # Write headers with styling
                header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                header_font = Font(bold=True, color="FFFFFF")
                
                for col_idx, header in enumerate(headers, 1):
                    cell = ws.cell(row=1, column=col_idx, value=header)
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                
                # Write data
                for row_idx, account in enumerate(accounts, 2):
                    # Same data as CSV
                    health_score = account.ai_health_score or 0
                    risk = 'Low' if health_score >= 80 else 'Medium' if health_score >= 50 else 'High'
                    
                    primary_contact_name = None
                    primary_contact_email = None
                    primary_contact_phone = None
                    try:
                        if account.primary_contact:
                            primary_contact_name = account.primary_contact.name
                            primary_contact_email = account.primary_contact.email
                            primary_contact_phone = account.primary_contact.phone
                    except Exception:
                        pass
                    
                    approval_status = "pending"
                    if account.account_approver and account.approval_date:
                        approval_status = "approved"
                    
                    row_data = [
                        account.client_name or '',
                        account.client_address.city if account.client_address else '',
                        account.hosting_area or '',
                        account.market_sector or '',
                        primary_contact_name or '',
                        account.client_type.value if account.client_type else '',
                        f"{health_score}%",
                        risk,
                        f"${(account.total_value or 0):.1f}M",
                        approval_status,
                        account.created_at.strftime('%Y-%m-%d') if account.created_at else '',
                        primary_contact_email or '',
                        primary_contact_phone or ''
                    ]
                    
                    for col_idx, value in enumerate(row_data, 1):
                        ws.cell(row=row_idx, column=col_idx, value=value)
                
                # Auto-adjust column widths
                for col in ws.columns:
                    max_length = 0
                    col_letter = col[0].column_letter
                    for cell in col:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)
                    ws.column_dimensions[col_letter].width = adjusted_width
                
                filename = f"accounts_export_{timestamp}.xlsx"
                temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
                temp_file.close()
                wb.save(temp_file.name)
                
                return FileResponse(
                    path=temp_file.name,
                    media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    filename=filename,
                    headers={"Content-Disposition": f"attachment; filename={filename}"}
                )
            except ImportError:
                logger.warning("openpyxl not installed, falling back to CSV")
                format = "csv"
        
        if format == "pdf":
            # PDF export (requires reportlab - install with: pip install reportlab)
            try:
                from reportlab.lib.pagesizes import letter
                from reportlab.lib import colors
                from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
                from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
                from reportlab.lib.units import inch
                
                filename = f"accounts_export_{timestamp}.pdf"
                temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.pdf')
                temp_file.close()
                
                doc = SimpleDocTemplate(temp_file.name, pagesize=letter)
                story = []
                
                # Title
                styles = getSampleStyleSheet()
                title_style = ParagraphStyle(
                    'CustomTitle',
                    parent=styles['Heading1'],
                    fontSize=16,
                    textColor=colors.HexColor('#366092'),
                    spaceAfter=30,
                )
                story.append(Paragraph("Accounts Export Report", title_style))
                story.append(Spacer(1, 0.2*inch))
                
                # Table data
                table_data = [headers]
                for account in accounts:
                    health_score = account.ai_health_score or 0
                    risk = 'Low' if health_score >= 80 else 'Medium' if health_score >= 50 else 'High'
                    
                    primary_contact_name = None
                    primary_contact_email = None
                    primary_contact_phone = None
                    try:
                        if account.primary_contact:
                            primary_contact_name = account.primary_contact.name
                            primary_contact_email = account.primary_contact.email
                            primary_contact_phone = account.primary_contact.phone
                    except Exception:
                        pass
                    
                    approval_status = "pending"
                    if account.account_approver and account.approval_date:
                        approval_status = "approved"
                    
                    table_data.append([
                        account.client_name or '',
                        account.client_address.city if account.client_address else '',
                        account.hosting_area or '',
                        account.market_sector or '',
                        primary_contact_name or '',
                        account.client_type.value if account.client_type else '',
                        f"{health_score}%",
                        risk,
                        f"${(account.total_value or 0):.1f}M",
                        approval_status,
                        account.created_at.strftime('%Y-%m-%d') if account.created_at else '',
                        primary_contact_email or '',
                        primary_contact_phone or ''
                    ])
                
                # Create table
                table = Table(table_data)
                table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#366092')),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                    ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, 0), 10),
                    ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                    ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                    ('GRID', (0, 0), (-1, -1), 1, colors.black),
                    ('FONTSIZE', (0, 1), (-1, -1), 8),
                ]))
                
                story.append(table)
                doc.build(story)
                
                return FileResponse(
                    path=temp_file.name,
                    media_type="application/pdf",
                    filename=filename,
                    headers={"Content-Disposition": f"attachment; filename={filename}"}
                )
            except ImportError:
                logger.warning("reportlab not installed, falling back to CSV")
                format = "csv"
        
        # Default CSV export
        filename = f"accounts_export_{timestamp}.csv"
        
        # Write to temporary file
        temp_file = tempfile.NamedTemporaryFile(mode='w', delete=False, suffix='.csv')
        temp_file.write(csv_content)
        temp_file.close()
        
        logger.info(f"Export successful, sending file: {filename}")
        
        # Return FileResponse
        return FileResponse(
            path=temp_file.name,
            media_type="text/csv",
            filename=filename,
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error exporting accounts: {str(e)}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"Failed to export accounts: {str(e)}")

@router.delete("/{account_id}/contacts/{contact_id}")
async def delete_account_contact(
    account_id: str,
    contact_id: str,
    current_user: AuthUserResponse = Depends(get_current_user)
):
    """Delete a contact from an account"""
    try:
        org_id = current_user.org_id
        if not org_id:
            raise HTTPException(status_code=401, detail="Organization ID not found")
        
        # Verify the account exists and belongs to the user's organization
        account = await account_service.get_account_by_id(UUID(account_id), org_id)
        if not account:
            raise HTTPException(status_code=404, detail="Account not found")
        
        # Get the contact and verify it belongs to the account
        from app.services.contact import ContactService
        contact_service = ContactService()
        contacts = await contact_service.get_contacts_by_account(UUID(account_id))
        
        contact_to_delete = None
        for contact in contacts:
            if str(contact.id) == contact_id:
                contact_to_delete = contact
                break
        
        if not contact_to_delete:
            raise HTTPException(status_code=404, detail="Contact not found")
        
        # Delete the contact
        from app.db.session import get_request_transaction
        db = get_request_transaction()
        
        await db.delete(contact_to_delete)
        await db.commit()
        
        logger.info(f"Deleted contact {contact_id} from account {account_id}")
        
        return {"message": "Contact deleted successfully"}
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error deleting contact {contact_id} from account {account_id}: {str(e)}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"Failed to delete contact: {str(e)}")

@router.get("/{account_id}", response_model=AccountDetailResponse)
async def get_account(
    account_id: str = Path(..., description="Account ID (UUID) or custom_id (e.g., AC-NY001) to retrieve"),
    current_user: AuthUserResponse = Depends(get_current_user)
):
    try:
        # Try to parse as UUID first, if it fails, treat as custom_id
        account = None
        try:
            account_uuid = UUID(account_id)
            account = await account_service.get_account_by_id(account_uuid, current_user.org_id)
        except (ValueError, TypeError):
            # Not a valid UUID, try as custom_id
            logger.info(f"Treating {account_id} as custom_id")
            account = await account_service.get_account_by_custom_id(account_id, current_user.org_id)
        
        if not account:
            raise HTTPException(status_code=404, detail="Account not found")
        
        # Build address data if available
        client_address = None
        if account.client_address:
            client_address = AddressResponse(
                address_id=account.client_address.id,
                line1=account.client_address.line1,
                line2=account.client_address.line2,
                city=account.client_address.city,
                state=account.client_address.state,
                pincode=account.client_address.pincode
            )
        
        # Build primary contact data if available
        primary_contact = None
        if account.primary_contact:
            primary_contact = ContactResponse(
                contact_id=account.primary_contact.id,
                name=account.primary_contact.name,
                email=account.primary_contact.email,
                phone=account.primary_contact.phone,
                title=account.primary_contact.title
            )
        
        # Build secondary contacts - load them separately to avoid lazy loading issues
        secondary_contacts = []
        try:
            # Load contacts separately to avoid lazy loading issues
            from app.models.contact import Contact
            from app.db.session import get_request_transaction
            
            db = get_request_transaction()
            contacts_stmt = select(Contact).where(Contact.account_id == account.account_id)
            contacts_result = await db.execute(contacts_stmt)
            account_contacts = contacts_result.scalars().all()
            
            for contact in account_contacts:
                if contact.id != account.primary_contact_id:  # Skip primary contact
                    secondary_contacts.append(ContactResponse(
                        contact_id=contact.id,
                        name=contact.name,
                        email=contact.email,
                        phone=contact.phone,
                        title=contact.title
                    ))
        except Exception as e:
            logger.error(f"Error loading secondary contacts: {str(e)}")
            # Continue without secondary contacts
        
        # Get creator and updater names
        created_by_name = None
        updated_by_name = None
        try:
            from app.models.user import User
            
            if account.created_by:
                creator_stmt = select(User).where(User.id == account.created_by)
                creator_result = await db.execute(creator_stmt)
                creator = creator_result.scalar_one_or_none()
                if creator:
                    created_by_name = creator.name or creator.email.split('@')[0] if creator.email else None
            
            if account.updated_by:
                updater_stmt = select(User).where(User.id == account.updated_by)
                updater_result = await db.execute(updater_stmt)
                updater = updater_result.scalar_one_or_none()
                if updater:
                    updated_by_name = updater.name or updater.email.split('@')[0] if updater.email else None
        except Exception as e:
            logger.error(f"Error loading creator/updater names: {str(e)}")
            # Continue without names
        
        return AccountDetailResponse(
            account_id=account.account_id,
            custom_id=account.custom_id,
            company_website=account.company_website,
            client_name=account.client_name,
            client_address=client_address,
            primary_contact=primary_contact,
            secondary_contacts=secondary_contacts,
            client_type=account.client_type,
            market_sector=account.market_sector,
            notes=account.notes,
            total_value=float(account.total_value) if account.total_value else None,
            ai_health_score=float(account.ai_health_score) if account.ai_health_score else None,
            health_trend=account.health_trend,
            risk_level=account.risk_level,
            last_ai_analysis=account.last_ai_analysis,
            data_quality_score=float(account.data_quality_score) if account.data_quality_score else None,
            revenue_growth=float(account.revenue_growth) if account.revenue_growth else None,
            communication_frequency=float(account.communication_frequency) if account.communication_frequency else None,
            win_rate=float(account.win_rate) if account.win_rate else None,
            opportunities=account.opportunities,
            last_contact=account.last_contact,
            hosting_area=account.hosting_area,
            account_approver=account.account_approver,
            approval_date=account.approval_date,
            created_at=account.created_at,
            updated_at=account.updated_at,
            created_by_name=created_by_name,
            updated_by_name=updated_by_name
        )
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error fetching account {account_id}: {str(e)}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"Failed to retrieve account: {str(e)}")

# Test endpoint without authentication
@router.get("/test", response_model=AccountListResponse)
async def test_accounts():
    return AccountListResponse(
        accounts=[],
        total=0,
        page=1,
        size=10,
        total_pages=0,
        pagination={
            "page": 1,
            "size": 10,
            "total": 0,
            "total_pages": 0
        }
    )

# Create new account
@router.post("/", response_model=AccountCreateResponse)
async def create_account(
    account_data: AccountCreate,
    current_user: AuthUserResponse = Depends(get_current_user),
    request: Request = None
):
    try:
        # Log the received data
        logger.info(f"Received account creation request for user {current_user.email}")
        
        # Create account using the service
        account = await account_service.create_account(
            org_id=current_user.org_id,
            account_data=account_data.dict() if hasattr(account_data, 'dict') else account_data,
            created_by=current_user.id
        )
        
        # Log audit trail
        try:
            from app.services.account_audit import account_audit_service
            db = get_request_transaction()
            await account_audit_service.log_account_change(
                db=db,
                account_id=account.account_id,
                action='create',
                user_id=current_user.id,
                summary=f"Account '{account.client_name}' created",
                request=request
            )
            await db.commit()
        except Exception as audit_error:
            logger.warning(f"Failed to log audit trail: {audit_error}")
        
        logger.info(f"Account created successfully: {account.account_id}")
        
        return AccountCreateResponse(
            account_id=str(account.account_id),
            client_name=account.client_name,
            message="Account created successfully"
        )
        
    except ValueError as e:
        logger.error(f"Validation error creating account: {str(e)}")
        raise HTTPException(status_code=400, detail=str(e))
    except Exception as e:
        logger.error(f"Error creating account: {str(e)}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"Failed to create account: {str(e)}")

# Approve account
@router.post("/{account_id}/approve", response_model=AccountUpdateResponse)
async def approve_account(
    request_data: AccountApprovalRequest,
    account_id: UUID = Path(..., description="Account ID to approve"),
    current_user: AuthUserResponse = Depends(get_current_user),
    request: Request = None
):
    try:
        # Get the account
        account = await account_service.get_account_by_id(account_id, current_user.org_id)
        if not account:
            raise HTTPException(status_code=404, detail="Account not found")
        
        # Update account status
        account.approval_status = "approved"
        if request_data.notes:
            account.approval_notes = request_data.notes
        account.account_approver = current_user.email
        account.approval_date = datetime.now()
        
        # Save to database
        db = get_request_transaction()
        db.add(account)
        await db.commit()
        
        # Log audit trail
        try:
            from app.services.account_audit import account_audit_service
            await account_audit_service.log_account_change(
                db=db,
                account_id=account.account_id,
                action='approve',
                user_id=current_user.id,
                summary=f"Account approved by {current_user.email}",
                request=request
            )
            await db.commit()
        except Exception as audit_error:
            logger.warning(f"Failed to log audit trail: {audit_error}")
        
        logger.info(f"Account {account_id} approved by {current_user.email}")
        
        return AccountUpdateResponse(
            account_id=str(account.account_id),
            message="Account approved successfully"
        )
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error approving account {account_id}: {str(e)}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"Failed to approve account: {str(e)}")

# Decline account
@router.post("/{account_id}/decline", response_model=AccountUpdateResponse)
async def decline_account(
    request_data: AccountApprovalRequest,
    account_id: UUID = Path(..., description="Account ID to decline"),
    current_user: AuthUserResponse = Depends(get_current_user),
    request: Request = None
):
    try:
        # Get the account
        account = await account_service.get_account_by_id(account_id, current_user.org_id)
        if not account:
            raise HTTPException(status_code=404, detail="Account not found")
        
        # Update account status
        account.approval_status = "declined"
        if request_data.notes:
            account.approval_notes = request_data.notes
        account.account_approver = current_user.email
        account.approval_date = datetime.now()
        
        # Save to database
        db = get_request_transaction()
        db.add(account)
        await db.commit()
        
        # Log audit trail
        try:
            from app.services.account_audit import account_audit_service
            await account_audit_service.log_account_change(
                db=db,
                account_id=account.account_id,
                action='decline',
                user_id=current_user.id,
                summary=f"Account declined by {current_user.email}",
                request=request
            )
            await db.commit()
        except Exception as audit_error:
            logger.warning(f"Failed to log audit trail: {audit_error}")
        
        logger.info(f"Account {account_id} declined by {current_user.email}")
        
        return AccountUpdateResponse(
            account_id=str(account.account_id),
            message="Account declined"
        )
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error declining account {account_id}: {str(e)}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"Failed to decline account: {str(e)}")

# Get contacts for an account
@router.get("/{account_id}/contacts", response_model=List[ContactResponse])
async def get_account_contacts(
    account_id: str,
    current_user: AuthUserResponse = Depends(get_current_user)
):
    try:
        from app.services.contact import ContactService
        
        # Verify the account exists and belongs to the user's organization
        account = await account_service.get_account_by_id(UUID(account_id), current_user.org_id)
        if not account:
            raise HTTPException(status_code=404, detail="Account not found")
        
        # Get contacts for the account
        contact_service = ContactService()
        contacts = await contact_service.get_contacts_by_account(UUID(account_id))
        
        # Convert to response format
        contact_responses = []
        for contact in contacts:
            contact_responses.append(ContactResponse(
                contact_id=contact.id,
                name=contact.name,
                email=contact.email,
                phone=contact.phone,
                title=contact.title
            ))
        
        logger.info(f"Retrieved {len(contact_responses)} contacts for account {account_id}")
        return contact_responses
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error getting contacts for account {account_id}: {str(e)}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"Failed to get contacts: {str(e)}")

# Add contact to an account
@router.post("/{account_id}/contacts", response_model=ContactResponse)
async def add_account_contact(
    account_id: str,
    contact_data: ContactCreate,
    current_user: AuthUserResponse = Depends(get_current_user),
    db: AsyncSession = Depends(get_request_transaction)
):
    try:
        logger.info(f"🔍 RAW contact_data type: {type(contact_data)}")
        logger.info(f"🔍 RAW contact_data: {contact_data}")
        logger.info(f"🔍 contact_data.model_dump(): {contact_data.model_dump()}")
        logger.info(f"🔍 Contact fields - name: '{contact_data.name}', email: '{contact_data.email}', phone: '{contact_data.phone}', title: '{contact_data.title}'")
        
        org_id = current_user.org_id
        if not org_id:
            raise HTTPException(status_code=401, detail="Organization ID not found")
        
        # Get the account
        # org_id is already a UUID from the database, no need to convert
        account = await account_service.get_account_by_id(UUID(account_id), org_id)
        if not account:
            raise HTTPException(status_code=404, detail="Account not found")
        
        # Create the contact
        from app.models.contact import Contact
        
        # Convert to UUID if needed
        account_uuid = UUID(account_id) if isinstance(account_id, str) else account_id
        # org_id is already a UUID from the database, no need to convert
        org_uuid = org_id
        
        contact = Contact(
            name=contact_data.name,
            email=contact_data.email,
            phone=contact_data.phone,
            title=contact_data.title,
            account_id=account_uuid,
            org_id=org_uuid
        )
        
        db.add(contact)
        await db.flush()
        
        # Get the ID from the flushed contact
        contact_id = contact.id
        
        # Query back to get the full contact data
        from app.services.contact import ContactService
        contact_service = ContactService()
        created_contact = await contact_service.get_contact_by_id(contact_id)
        
        if not created_contact:
            raise HTTPException(status_code=500, detail="Failed to retrieve created contact")
        
        logger.info(f"Successfully created contact {created_contact.id} for account {account_id}")
        
        return ContactResponse(
            contact_id=created_contact.id,
            name=created_contact.name,
            email=created_contact.email,
            phone=created_contact.phone,
            title=created_contact.title
        )
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error adding contact to account {account_id}: {str(e)}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"Failed to add contact: {str(e)}")

# Delete contact from an account
@router.put("/{account_id}/contacts/{contact_id}", response_model=ContactResponse)
async def update_account_contact(
    account_id: str,
    contact_id: str,
    contact_data: ContactUpdateRequest,
    current_user: AuthUserResponse = Depends(get_current_user)
):
    """Update a contact for an account"""
    try:
        org_id = current_user.org_id
        if not org_id:
            raise HTTPException(status_code=401, detail="Organization ID not found")
        
        # Verify the account exists and belongs to the user's organization
        account = await account_service.get_account_by_id(UUID(account_id), org_id)
        if not account:
            raise HTTPException(status_code=404, detail="Account not found")
        
        # Update contact
        from app.services.contact import ContactService
        contact_service = ContactService()
        updated_contact = await contact_service.update_contact(
            UUID(contact_id),
            contact_data.model_dump(exclude_unset=True)
        )
        
        if not updated_contact:
            raise HTTPException(status_code=404, detail="Contact not found")
        
        logger.info(f"Successfully updated contact {contact_id} for account {account_id}")
            
        return ContactResponse(
            contact_id=updated_contact.id,
            name=updated_contact.name,
            email=updated_contact.email,
            phone=updated_contact.phone,
            title=updated_contact.title
        )
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error updating contact {contact_id}: {str(e)}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"Failed to update contact: {str(e)}")

# Update account
@router.put("/{account_id}", response_model=AccountUpdateResponse)
async def update_account(
    account_id: UUID,
    account_data: AccountUpdate,
    current_user: AuthUserResponse = Depends(get_current_user),
    request: Request = None
):
    """Update account details including address and primary contact"""
    try:
        # Get the account
        account = await account_service.get_account_by_id(account_id, current_user.org_id)
        if not account:
            raise HTTPException(status_code=404, detail="Account not found")
        
        # Store old data for audit trail
        old_data = {
            'client_name': account.client_name,
            'client_type': account.client_type.value if account.client_type else None,
            'market_sector': account.market_sector,
            'company_website': account.company_website,
            'hosting_area': account.hosting_area,
            'notes': account.notes,
        }
        
        # Update account fields
        if account_data.client_name is not None:
            account.client_name = account_data.client_name
        if account_data.client_type is not None:
            account.client_type = account_data.client_type
        if account_data.market_sector is not None:
            account.market_sector = account_data.market_sector
        if account_data.company_website is not None:
            account.company_website = str(account_data.company_website) if account_data.company_website else None
        if account_data.hosting_area is not None:
            account.hosting_area = account_data.hosting_area
        if account_data.notes is not None:
            account.notes = account_data.notes
        if account_data.account_approver is not None:
            account.account_approver = account_data.account_approver
        if account_data.approval_date is not None:
            account.approval_date = account_data.approval_date
        
        # Update address if provided
        if account_data.client_address is not None:
            from app.models.address import Address
            
            if account.client_address:
                # Update existing address
                if account_data.client_address.line1:
                    account.client_address.line1 = account_data.client_address.line1
                if account_data.client_address.line2 is not None:
                    account.client_address.line2 = account_data.client_address.line2
                if account_data.client_address.city is not None:
                    account.client_address.city = account_data.client_address.city
                if account_data.client_address.state is not None:
                    account.client_address.state = account_data.client_address.state
                if account_data.client_address.pincode is not None:
                    account.client_address.pincode = account_data.client_address.pincode
            elif account_data.client_address.line1:
                # Create new address
                new_address = Address(
                    line1=account_data.client_address.line1,
                    line2=account_data.client_address.line2,
                    city=account_data.client_address.city,
                    state=account_data.client_address.state,
                    pincode=account_data.client_address.pincode,
                    org_id=current_user.org_id
                )
                db = get_request_transaction()
                db.add(new_address)
                await db.flush()
                account.client_address_id = new_address.id
        
        # Set updated_by
        account.updated_by = current_user.id
        
        # Save to database
        db = get_request_transaction()
        db.add(account)
        await db.commit()
        
        # Log audit trail
        try:
            from app.services.account_audit import account_audit_service
            new_data = {
                'client_name': account.client_name,
                'client_type': account.client_type.value if account.client_type else None,
                'market_sector': account.market_sector,
                'company_website': account.company_website,
                'hosting_area': account.hosting_area,
                'notes': account.notes,
            }
            changes = account_audit_service.compute_field_changes(old_data, new_data)
            summary = account_audit_service.generate_summary('update', changes)
            
            await account_audit_service.log_account_change(
                db=db,
                account_id=account.account_id,
                action='update',
                user_id=current_user.id,
                changes=changes,
                summary=summary,
                request=request
            )
            await db.commit()
        except Exception as audit_error:
            logger.warning(f"Failed to log audit trail: {audit_error}")
        
        logger.info(f"Account {account_id} updated by {current_user.email}")
        
        return AccountUpdateResponse(
            account_id=str(account.account_id),
            message="Account updated successfully"
        )
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error updating account {account_id}: {str(e)}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"Failed to update account: {str(e)}")

# Get account audit trail
@router.get("/{account_id}/audit-trail")
async def get_account_audit_trail(
    account_id: UUID = Path(..., description="Account ID"),
    limit: int = Query(50, ge=1, le=200, description="Maximum number of audit logs to return"),
    action: Optional[str] = Query(None, description="Filter by action type"),
    current_user: AuthUserResponse = Depends(get_current_user)
):
    """Get audit trail (change history) for an account"""
    try:
        # Verify account exists and user has access
        account = await account_service.get_account_by_id(account_id, current_user.org_id)
        if not account:
            raise HTTPException(status_code=404, detail="Account not found")
        
        # Get audit logs
        from app.services.account_audit import account_audit_service
        db = get_request_transaction()
        audit_logs = await account_audit_service.get_account_audit_history(
            db=db,
            account_id=account_id,
            limit=limit,
            action_filter=action
        )
        
        return {
            "account_id": str(account_id),
            "total_logs": len(audit_logs),
            "logs": [log.to_dict() for log in audit_logs]
        }
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error fetching audit trail for account {account_id}: {str(e)}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"Failed to fetch audit trail: {str(e)}")

# Bulk operations
@router.post("/bulk/delete")
async def bulk_delete_accounts(
    account_ids: List[UUID],
    current_user: AuthUserResponse = Depends(get_current_user),
    request: Request = None
):
    """Bulk delete accounts"""
    try:
        if not account_ids:
            raise HTTPException(status_code=400, detail="No account IDs provided")
        
        db = get_request_transaction()
        deleted_count = 0
        
        for account_id in account_ids:
            account = await account_service.get_account_by_id(account_id, current_user.org_id)
            if account:
                # Log audit trail before deletion
                try:
                    from app.services.account_audit import account_audit_service
                    await account_audit_service.log_account_change(
                        db=db,
                        account_id=account.account_id,
                        action='delete',
                        user_id=current_user.id,
                        summary=f"Account '{account.client_name}' deleted",
                        request=request
                    )
                except Exception as audit_error:
                    logger.warning(f"Failed to log audit trail: {audit_error}")
                
                await db.delete(account)
                deleted_count += 1
        
        await db.commit()
        
        logger.info(f"Bulk deleted {deleted_count} accounts by {current_user.email}")
        
        return {
            "message": f"Successfully deleted {deleted_count} account(s)",
            "deleted_count": deleted_count,
            "total_requested": len(account_ids)
        }
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error in bulk delete: {str(e)}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"Failed to delete accounts: {str(e)}")

@router.post("/bulk/update")
async def bulk_update_accounts(
    updates: List[dict],  # List of {account_id, updates}
    current_user: AuthUserResponse = Depends(get_current_user),
    request: Request = None
):
    """Bulk update accounts"""
    try:
        if not updates:
            raise HTTPException(status_code=400, detail="No updates provided")
        
        db = get_request_transaction()
        updated_count = 0
        
        for update_item in updates:
            account_id = UUID(update_item.get('account_id'))
            update_data = update_item.get('updates', {})
            
            account = await account_service.get_account_by_id(account_id, current_user.org_id)
            if account:
                # Store old data for audit
                old_data = {
                    'client_name': account.client_name,
                    'client_type': account.client_type.value if account.client_type else None,
                    'market_sector': account.market_sector,
                }
                
                # Apply updates
                if 'client_name' in update_data:
                    account.client_name = update_data['client_name']
                if 'market_sector' in update_data:
                    account.market_sector = update_data['market_sector']
                if 'hosting_area' in update_data:
                    account.hosting_area = update_data['hosting_area']
                
                account.updated_by = current_user.id
                db.add(account)
                
                # Log audit trail
                try:
                    from app.services.account_audit import account_audit_service
                    new_data = {
                        'client_name': account.client_name,
                        'client_type': account.client_type.value if account.client_type else None,
                        'market_sector': account.market_sector,
                    }
                    changes = account_audit_service.compute_field_changes(old_data, new_data)
                    summary = account_audit_service.generate_summary('update', changes)
                    
                    await account_audit_service.log_account_change(
                        db=db,
                        account_id=account.account_id,
                        action='update',
                        user_id=current_user.id,
                        changes=changes,
                        summary=f"Bulk update: {summary}",
                        request=request
                    )
                except Exception as audit_error:
                    logger.warning(f"Failed to log audit trail: {audit_error}")
                
                updated_count += 1
        
        await db.commit()
        
        logger.info(f"Bulk updated {updated_count} accounts by {current_user.email}")
        
        return {
            "message": f"Successfully updated {updated_count} account(s)",
            "updated_count": updated_count,
            "total_requested": len(updates)
        }
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error in bulk update: {str(e)}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"Failed to update accounts: {str(e)}")

# Soft delete (archive) account
@router.post("/{account_id}/archive")
async def archive_account(
    account_id: UUID = Path(..., description="Account ID to archive"),
    current_user: AuthUserResponse = Depends(get_current_user),
    request: Request = None
):
    """Soft delete (archive) an account"""
    try:
        account = await account_service.get_account_by_id(account_id, current_user.org_id)
        if not account:
            raise HTTPException(status_code=404, detail="Account not found")
        
        if account.is_deleted:
            raise HTTPException(status_code=400, detail="Account is already archived")
        
        # Soft delete
        account.is_deleted = True
        account.deleted_at = datetime.now()
        account.deleted_by = current_user.id
        
        db = get_request_transaction()
        db.add(account)
        await db.commit()
        
        # Log audit trail
        try:
            from app.services.account_audit import account_audit_service
            await account_audit_service.log_account_change(
                db=db,
                account_id=account.account_id,
                action='delete',
                user_id=current_user.id,
                summary=f"Account '{account.client_name}' archived",
                request=request
            )
            await db.commit()
        except Exception as audit_error:
            logger.warning(f"Failed to log audit trail: {audit_error}")
        
        logger.info(f"Account {account_id} archived by {current_user.email}")
        
        return AccountUpdateResponse(
            account_id=str(account.account_id),
            message="Account archived successfully"
        )
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error archiving account {account_id}: {str(e)}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"Failed to archive account: {str(e)}")

# Restore archived account
@router.post("/{account_id}/restore")
async def restore_account(
    account_id: UUID = Path(..., description="Account ID to restore"),
    current_user: AuthUserResponse = Depends(get_current_user),
    request: Request = None
):
    """Restore an archived account"""
    try:
        db = get_request_transaction()
        
        # Get account including archived ones
        stmt = select(Account).where(
            and_(
                Account.account_id == account_id,
                Account.org_id == current_user.org_id
            )
        )
        result = await db.execute(stmt)
        account = result.scalar_one_or_none()
        
        if not account:
            raise HTTPException(status_code=404, detail="Account not found")
        
        if not account.is_deleted:
            raise HTTPException(status_code=400, detail="Account is not archived")
        
        # Restore
        account.is_deleted = False
        account.deleted_at = None
        account.deleted_by = None
        account.updated_by = current_user.id
        
        db.add(account)
        await db.commit()
        
        # Log audit trail
        try:
            from app.services.account_audit import account_audit_service
            await account_audit_service.log_account_change(
                db=db,
                account_id=account.account_id,
                action='update',
                user_id=current_user.id,
                summary=f"Account '{account.client_name}' restored from archive",
                request=request
            )
            await db.commit()
        except Exception as audit_error:
            logger.warning(f"Failed to log audit trail: {audit_error}")
        
        logger.info(f"Account {account_id} restored by {current_user.email}")
        
        return AccountUpdateResponse(
            account_id=str(account.account_id),
            message="Account restored successfully"
        )
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error restoring account {account_id}: {str(e)}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"Failed to restore account: {str(e)}")

# Get archived accounts
@router.get("/archived", response_model=AccountListResponse)
async def get_archived_accounts(
    page: int = Query(1, ge=1),
    size: int = Query(10, ge=1, le=100),
    current_user: AuthUserResponse = Depends(get_current_user)
):
    """Get archived (soft-deleted) accounts"""
    try:
        db = get_request_transaction()
        
        stmt = (
            select(Account)
            .where(
                and_(
                    Account.org_id == current_user.org_id,
                    Account.is_deleted == True
                )
            )
            .order_by(Account.deleted_at.desc())
            .offset((page - 1) * size)
            .limit(size)
        )
        
        result = await db.execute(stmt)
        accounts = list(result.scalars().all())
        
        # Get total count
        count_stmt = select(func.count(Account.account_id)).where(
            and_(
                Account.org_id == current_user.org_id,
                Account.is_deleted == True
            )
        )
        count_result = await db.execute(count_stmt)
        total = count_result.scalar() or 0
        
        # Convert to response format
        account_items = []
        for account in accounts:
            primary_contact_name = None
            primary_contact_email = None
            try:
                if account.primary_contact:
                    primary_contact_name = account.primary_contact.name
                    primary_contact_email = account.primary_contact.email
            except Exception:
                pass
            
            account_items.append(AccountListItem(
                account_id=str(account.account_id),
                custom_id=account.custom_id,
                client_name=account.client_name,
                client_type=account.client_type.value if account.client_type else None,
                market_sector=account.market_sector,
                hosting_area=account.hosting_area,
                total_value=float(account.total_value) if account.total_value else 0.0,
                ai_health_score=float(account.ai_health_score) if account.ai_health_score else None,
                health_trend=account.health_trend,
                risk_level=account.risk_level,
                approval_status=account.approval_status,
                created_at=account.created_at,
            ))
        
        return AccountListResponse(
            accounts=account_items,
            pagination={
                "total": total,
                "page": page,
                "size": size,
                "total_pages": (total + size - 1) // size if total > 0 else 0,
                "has_next": (page * size) < total,
                "has_prev": page > 1
            }
        )
        
    except Exception as e:
        logger.error(f"Error fetching archived accounts: {str(e)}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"Failed to fetch archived accounts: {str(e)}")

# Advanced Analytics endpoints
@router.get("/analytics/revenue")
async def get_revenue_analytics(
    time_period: str = Query("30d", regex="^(7d|30d|90d|1y|all)$", description="Time period for analytics"),
    account_id: Optional[UUID] = Query(None, description="Filter by specific account"),
    current_user: AuthUserResponse = Depends(get_current_user)
):
    """Get revenue analytics with trends"""
    try:
        from app.services.account_analytics import account_analytics_service
        
        analytics = await account_analytics_service.get_revenue_analytics(
            org_id=current_user.org_id,
            time_period=time_period,
            account_id=account_id
        )
        
        return analytics
        
    except Exception as e:
        logger.error(f"Error fetching revenue analytics: {str(e)}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"Failed to fetch revenue analytics: {str(e)}")

@router.get("/analytics/performance")
async def get_performance_dashboard(
    account_id: Optional[UUID] = Query(None, description="Filter by specific account"),
    current_user: AuthUserResponse = Depends(get_current_user)
):
    """Get performance dashboard with KPIs"""
    try:
        from app.services.account_analytics import account_analytics_service
        
        dashboard = await account_analytics_service.get_performance_dashboard(
            org_id=current_user.org_id,
            account_id=account_id
        )
        
        return dashboard
        
    except Exception as e:
        logger.error(f"Error fetching performance dashboard: {str(e)}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"Failed to fetch performance dashboard: {str(e)}")

@router.get("/analytics/growth")
async def get_growth_tracking(
    account_id: Optional[UUID] = Query(None, description="Filter by specific account"),
    months: int = Query(12, ge=1, le=24, description="Number of months to track"),
    current_user: AuthUserResponse = Depends(get_current_user)
):
    """Get growth tracking over time"""
    try:
        from app.services.account_analytics import account_analytics_service
        
        growth = await account_analytics_service.get_growth_tracking(
            org_id=current_user.org_id,
            account_id=account_id,
            months=months
        )
        
        return growth
        
    except Exception as e:
        logger.error(f"Error fetching growth tracking: {str(e)}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"Failed to fetch growth tracking: {str(e)}")

@router.post("/analytics/compare")
async def compare_accounts(
    account_ids: List[UUID],
    current_user: AuthUserResponse = Depends(get_current_user)
):
    """Compare multiple accounts side by side"""
    try:
        if not account_ids or len(account_ids) < 2:
            raise HTTPException(status_code=400, detail="At least 2 account IDs required for comparison")
        
        if len(account_ids) > 10:
            raise HTTPException(status_code=400, detail="Maximum 10 accounts can be compared at once")
        
        from app.services.account_analytics import account_analytics_service
        
        comparison = await account_analytics_service.compare_accounts(
            account_ids=account_ids,
            org_id=current_user.org_id
        )
        
        return comparison
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error comparing accounts: {str(e)}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"Failed to compare accounts: {str(e)}")

# Project Sheet Generation endpoints
@router.post("/{account_id}/project-sheet")
async def generate_project_sheet(
    account_id: UUID = Path(..., description="Account ID"),
    template_type: str = Query("comprehensive", regex="^(comprehensive|executive|technical|financial)$", description="Template type"),
    include_history: bool = Query(True, description="Include historical data"),
    current_user: AuthUserResponse = Depends(get_current_user)
):
    """Generate AI-powered project sheet for an account"""
    try:
        from app.services.project_sheet_generation import project_sheet_generation_service
        
        sheet = await project_sheet_generation_service.generate_project_sheet(
            account_id=account_id,
            org_id=current_user.org_id,
            template_type=template_type,
            include_history=include_history
        )
        
        return sheet
        
    except ValueError as e:
        raise HTTPException(status_code=404, detail=str(e))
    except Exception as e:
        logger.error(f"Error generating project sheet: {str(e)}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"Failed to generate project sheet: {str(e)}")

# Health check endpoint
@router.get("/health", status_code=200)
async def health_check():
    return {"status": "ok", "message": "Account routes are working"}